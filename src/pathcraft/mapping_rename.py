import csv
import os
from collections.abc import Iterable
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from .config import MAPPING_EXTENSIONS
from .filesystem import file_signature
from .rename import RenameEntry, mark_existing_destination_conflicts
from .scanner import iter_files
from .utils import filename_validation_error


SUPPORTED_MAPPING_EXTENSIONS = MAPPING_EXTENSIONS


class _SemicolonDialect(csv.excel):
    delimiter = ";"


class _PipeDialect(csv.excel):
    delimiter = "|"


_FALLBACK_DIALECTS: dict[str, type[csv.Dialect]] = {
    ",": csv.excel,
    "\t": csv.excel_tab,
    ";": _SemicolonDialect,
    "|": _PipeDialect,
}


@dataclass(frozen=True)
class MappingTable:
    columns: tuple[str, ...]
    rows: tuple[dict[str, str], ...]

    def mappings(self, source_column: str, destination_column: str) -> list[tuple[str, str]]:
        if source_column not in self.columns:
            raise ValueError(f"未找到原名称列：{source_column}")
        if destination_column not in self.columns:
            raise ValueError(f"未找到新名称列：{destination_column}")

        mappings = []
        for row_number, row in enumerate(self.rows, start=2):
            source = row[source_column].strip()
            destination = row[destination_column].strip()
            if not source and not destination:
                continue
            if not source:
                raise ValueError(f"第 {row_number} 行的原名称为空")
            if not destination:
                raise ValueError(f"第 {row_number} 行的新名称为空")
            mappings.append((source, destination))
        if not mappings:
            raise ValueError("映射表中没有可用的名称映射")
        return mappings


def load_mapping_table(path: Path) -> MappingTable:
    suffix = path.suffix.lower()
    if suffix not in SUPPORTED_MAPPING_EXTENSIONS:
        supported = "、".join(sorted(SUPPORTED_MAPPING_EXTENSIONS))
        raise ValueError(f"不支持的映射文件格式：{suffix}（支持 {supported}）")
    if suffix in {".xlsx", ".xlsm"}:
        return _load_excel_table(path)
    return _load_text_table(path)


def load_mapping_columns(path: Path) -> tuple[str, ...]:
    suffix = path.suffix.lower()
    if suffix not in SUPPORTED_MAPPING_EXTENSIONS:
        supported = "、".join(sorted(SUPPORTED_MAPPING_EXTENSIONS))
        raise ValueError(f"不支持的映射文件格式：{suffix}（支持 {supported}）")
    if suffix in {".xlsx", ".xlsm"}:
        return _load_excel_columns(path)
    return _load_text_columns(path)


def _load_text_columns(path: Path) -> tuple[str, ...]:
    errors = []
    for encoding in ("utf-8-sig", "gb18030"):
        try:
            with path.open("r", encoding=encoding, newline="") as stream:
                sample = stream.read(8192)
                if not sample.strip():
                    raise ValueError("映射文件为空")
                stream.seek(0)
                dialect = _detect_dialect(sample, path.suffix.lower())
                header = next(csv.reader(stream, dialect), None)
        except UnicodeDecodeError as error:
            errors.append(error)
            continue
        if header is None:
            raise ValueError("映射文件为空")
        return _columns_from_values(header)
    raise ValueError(f"无法识别映射文件编码：{errors[-1]}")


def _load_text_table(path: Path) -> MappingTable:
    errors = []
    for encoding in ("utf-8-sig", "gb18030"):
        try:
            with path.open("r", encoding=encoding, newline="") as stream:
                sample = stream.read(8192)
                if not sample.strip():
                    raise ValueError("映射文件为空")
                stream.seek(0)
                dialect = _detect_dialect(sample, path.suffix.lower())
                return _table_from_rows(csv.reader(stream, dialect))
        except UnicodeDecodeError as error:
            errors.append(error)
    raise ValueError(f"无法识别映射文件编码：{errors[-1]}")


def _detect_dialect(sample: str, suffix: str) -> type[csv.Dialect] | csv.Dialect:
    if not sample.strip():
        return csv.excel
    try:
        return csv.Sniffer().sniff(sample, delimiters=",\t;|")
    except csv.Error:
        first_line = next((line for line in sample.splitlines() if line.strip()), "")
        delimiter = max(_FALLBACK_DIALECTS, key=first_line.count)
        if first_line.count(delimiter):
            return _FALLBACK_DIALECTS[delimiter]
        return csv.excel_tab if suffix == ".txt" else csv.excel


def _load_excel_columns(path: Path) -> tuple[str, ...]:
    try:
        import openpyxl
    except ImportError as error:
        raise ValueError("读取 Excel 需要 openpyxl，请先运行 uv sync") from error

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        header = next(workbook.active.iter_rows(values_only=True), None)
    finally:
        workbook.close()
    if header is None:
        raise ValueError("Excel 映射表为空")
    return _columns_from_values(header)


def _load_excel_table(path: Path) -> MappingTable:
    try:
        import openpyxl
    except ImportError as error:
        raise ValueError("读取 Excel 需要 openpyxl，请先运行 uv sync") from error

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        return _table_from_rows(
            worksheet.iter_rows(values_only=True),
            empty_message="Excel 映射表为空",
        )
    finally:
        workbook.close()


def _table_from_rows(
    raw_rows: Iterable[Iterable[Any]],
    *,
    empty_message: str = "映射文件为空",
) -> MappingTable:
    row_iter = iter(raw_rows)
    header = next(row_iter, None)
    if header is None:
        raise ValueError(empty_message)
    columns = _columns_from_values(tuple(header))

    rows = []
    for raw_values in row_iter:
        values = tuple(raw_values)
        padded = [*values, *([None] * (len(columns) - len(values)))]
        rows.append(
            {
                column: _cell_text(value)
                for column, value in zip(columns, padded[: len(columns)])
            }
        )
    return MappingTable(columns, tuple(rows))


def _columns_from_values(values: list[Any] | tuple[Any, ...]) -> tuple[str, ...]:
    columns = tuple(_cell_text(value) for value in values)
    if len(columns) < 2:
        raise ValueError("映射表至少需要两列")
    if any(not column for column in columns):
        raise ValueError("映射表包含空列标题")
    if len(set(columns)) != len(columns):
        raise ValueError("映射表包含重复的列标题")
    return columns


def _cell_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def build_mapping_plan(
    root: Path,
    mappings: list[tuple[str, str]],
    recursive: bool = True,
) -> list[RenameEntry]:
    relative_keys: set[str] = set()
    name_keys: set[str] = set()
    stem_keys: set[str] = set()
    for source_text, _ in mappings:
        normalized = source_text.replace("\\", "/").strip().strip("/")
        source_path = Path(normalized)
        if source_path.is_absolute() or ".." in source_path.parts:
            continue
        key = normalized.casefold()
        if "/" in normalized:
            relative_keys.add(key)
        else:
            name_keys.add(key)
            if not source_path.suffix:
                stem_keys.add(key)

    by_relative = {key: [] for key in relative_keys}
    by_name = {key: [] for key in name_keys}
    by_stem = {key: [] for key in stem_keys}
    for path in iter_files(root, recursive=recursive):
        if relative_keys:
            relative = path.relative_to(root).as_posix().casefold()
            if relative in by_relative:
                by_relative[relative].append(path)
        name = path.name.casefold()
        if name in by_name:
            by_name[name].append(path)
        if stem_keys:
            stem = path.stem.casefold()
            if stem in by_stem:
                by_stem[stem].append(path)

    preliminary = []
    used_sources: set[Path] = set()
    for source_text, destination_text in mappings:
        normalized_source = source_text.replace("\\", "/").strip().strip("/")
        source_relative = Path(normalized_source)
        if source_relative.is_absolute() or ".." in source_relative.parts:
            preliminary.append(
                RenameEntry(root / source_relative.name, root / destination_text, "原名称路径无效")
            )
            continue

        if "/" in normalized_source:
            candidates = by_relative.get(normalized_source.casefold(), [])
        else:
            candidates = list(by_name.get(normalized_source.casefold(), []))
            if not Path(normalized_source).suffix:
                candidates.extend(by_stem.get(normalized_source.casefold(), []))
            candidates = list(dict.fromkeys(candidates))

        placeholder = root / normalized_source
        if not candidates:
            preliminary.append(RenameEntry(placeholder, placeholder, "未找到原文件"))
            continue
        if len(candidates) > 1:
            preliminary.append(RenameEntry(placeholder, placeholder, "原名称匹配多个文件"))
            continue

        source = candidates[0]
        if source in used_sources:
            preliminary.append(RenameEntry(source, source, "原名称在映射表中重复"))
            continue
        used_sources.add(source)

        destination_name = destination_text.strip()
        if Path(destination_name).name != destination_name or "/" in destination_name or "\\" in destination_name:
            preliminary.append(RenameEntry(source, source, "新名称只能是文件名，不能包含目录"))
            continue
        if not Path(destination_name).suffix:
            destination_name = f"{destination_name}{source.suffix}"
        destination = source.with_name(destination_name)
        problem = filename_validation_error(destination.name)
        if source == destination:
            problem = "名称未变化"
        signature = None
        try:
            signature = file_signature(source)
        except OSError as error:
            if problem is None:
                problem = f"无法读取源文件：{error}"
        preliminary.append(RenameEntry(source, destination, problem, signature))

    destination_counts: dict[str, int] = {}
    for entry in preliminary:
        if entry.problem is None:
            key = os.path.normcase(str(entry.destination.absolute()))
            destination_counts[key] = destination_counts.get(key, 0) + 1
    plan = [
        RenameEntry(
            entry.source,
            entry.destination,
            "目标名称重复",
            entry.source_signature,
        )
        if entry.problem is None
        and destination_counts[os.path.normcase(str(entry.destination.absolute()))] > 1
        else entry
        for entry in preliminary
    ]
    return mark_existing_destination_conflicts(plan)
